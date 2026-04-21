"""
엑셀 -> data.js 변환 스크립트.

사용법:
    python build_data.py

동작:
    - 같은 폴더의 '부산_맛집_리서치_v*.xlsx' 중 가장 최신(파일명 기준) 하나를 읽음
    - 맛집_사진/ 폴더의 'NN_' 숫자 접두사로 행 순서에 맞춰 사진을 자동 매칭
    - data.js 파일을 생성 (busan.html이 <script src="data.js">로 로드)

엑셀을 수정하면 이 스크립트를 다시 돌리기만 하면 됩니다.
"""

import json
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
PHOTO_DIR = "맛집_사진"

candidates = sorted(BASE.glob("부산_맛집_리서치_v*.xlsx"))
if not candidates:
    candidates = sorted(BASE.glob("부산_맛집_리서치.xlsx"))
if not candidates:
    sys.exit("ERR: 부산_맛집_리서치*.xlsx 를 찾을 수 없습니다.")

XLSX = candidates[-1]
print(f"Source: {XLSX.name}", file=sys.stderr)

COL_MAP = {
    "지역": "region",
    "가게명": "name",
    "평점": "rating",
    "평점출처": "src",
    "카카오평점": "kakao",
    "주소": "addr",
    "대표메뉴(대표가격)": "menu",
    "웨이팅/예약": "wait",
    "칭찬포인트": "praise",
    "불만포인트": "complaint",
    "영업시간/휴무": "hours",
    "원본메모": "memo",
    "참고URL": "url",
}


def as_str(v):
    if v is None:
        return ""
    return str(v).strip()


def as_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip()
    if t in ("", "-"):
        return None
    try:
        return float(t)
    except ValueError:
        return None


photo_map = {}
pdir = BASE / PHOTO_DIR
if pdir.exists():
    for p in sorted(pdir.iterdir()):
        name = p.name
        if len(name) >= 3 and name[:2].isdigit() and name[2] == "_":
            photo_map[int(name[:2])] = f"{PHOTO_DIR}/{name}"

wb = openpyxl.load_workbook(XLSX)


def sheet_to_list(ws, start_idx):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [as_str(c) for c in rows[0]]
    out = []
    idx = start_idx
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {}
        for col, val in zip(header, row):
            key = COL_MAP.get(col)
            if not key:
                continue
            rec[key] = as_num(val) if key in ("rating", "kakao") else as_str(val)
        rec["photo"] = photo_map.get(idx, "")
        out.append(rec)
        idx += 1
    return out, idx


meal, next_idx = ([], 1)
if "식사" in wb.sheetnames:
    meal, next_idx = sheet_to_list(wb["식사"], 1)

bakery = []
if "빵·디저트" in wb.sheetnames:
    bakery, _ = sheet_to_list(wb["빵·디저트"], next_idx)


def dump(data):
    return json.dumps(data, ensure_ascii=False, indent=2)


content = (
    "// 자동 생성됨 -- build_data.py가 생성. 직접 편집하지 마세요.\n"
    f"// Source: {XLSX.name}\n"
    "// 엑셀 수정 후 `python build_data.py`로 재생성.\n\n"
    f"const MEAL = {dump(meal)};\n\n"
    f"const BAKERY = {dump(bakery)};\n"
)

out = BASE / "data.js"
out.write_text(content, encoding="utf-8")
print(
    f"Written: {out.name}  (MEAL: {len(meal)}, BAKERY: {len(bakery)})",
    file=sys.stderr,
)
